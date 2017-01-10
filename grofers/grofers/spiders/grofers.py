import scrapy
from scrapy.selector import Selector
from scrapy.spider import BaseSpider
from xlrd import open_workbook
from openpyxl import Workbook, load_workbook
import xlrd,sys
import openpyxl,os
from selenium import webdriver
from scrapy.contrib.spiders import CrawlSpider, Rule
from scrapy.contrib.linkextractors.sgml import SgmlLinkExtractor
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
#from selenium.webdriver.firefox.firefox_binary import FirefoxBinary

class linkedInSpider(scrapy.Spider):
	name = "grofers"
	allowed_domains = ["grofers.com"]

	urls = []
	links = []

	wb = openpyxl.load_workbook('../../../media/Benchmarking.xlsx')
	ws = wb.get_sheet_by_name('Done')
	
	list2 = []
	for row in range(2, ws.max_row + 1):
		list = []
		# Each row in the spreadsheet has data for one census tract.
		ref  = ws['A' + str(row)].value
		link = ws['R' + str(row)].value
		#print ref,link
		if link and (link.encode('utf8')).startswith("https:"):
			list.append(ref)
			list.append(link)
			list2.append(list) 
	
	#print list2

	for ref,link in list2:
		links.append(link)
	print links
	start_urls = links#["https://grofers.com/prn/brooke-bond-red-label-zip-lock-tea/prid/104234"]#["https://grofers.com/prn/parle-hide-seek-chocolate-chip-cookie/prid/11102"]

	#rules = (
	#Rule(SgmlLinkExtractor(), callback='parse_item', follow=True),
	#)
	
        #driver = webdriver.Firefox()
	driver = webdriver.Chrome(executable_path="../../../media/chromedriver")

	def parse(self, response):
		reload(sys)
		sys.setdefaultencoding('utf-8')
		import base64
		import csv
        	import cStringIO
		self.driver.get(response.url)
        	self.driver.implicitly_wait(20)
		#sel = Selector(response)
		#hxs = HtmlXPathSelector(text=response.body)
		
		print "!!!!!!!!!!!!!!!!!!!!"
		#names = sel.xpath('//div[@class="pdp-product__price--new"]/span/text()').extract()
		names = self.driver.find_element_by_xpath('//h1[@class="pdp-product__name weight--light"]/div').text
		#names = WebDriverWait(self.driver, 60).until(EC.presence_of_element_located((By.XPATH, 'h1[@class="pdpproduct__name weight--light"]/div')))
		name = names#.text
		#print "Name",name
		prod_name = []
		prod_price = []
		"""
		if len(names)>1:
			l = 0
			for name in names:
				l += 1
				product_name = (((name).strip().rstrip()).encode('utf-8')).replace("\\s+","")
				prod_name.append(product_name)
				if l < len(names):
					prod_name.append("\n")
		else:
			product_name = (((names[0]).strip().rstrip()).encode('utf-8')).replace("\\s+"," ")
			prod_name.append(product_name)
		product_name = (''.join(prod_name)).strip().rstrip()
		"""
	
		prices = self.driver.find_element_by_xpath('//div[@class="pdp-product__price"]/span').text
		#prices = WebDriverWait(self.driver, 60).until(EC.presence_of_element_located((By.XPATH, '//div[@class="pdp-product__price"]/span')))
		price = prices#.text
		print "Prices",price
		
		"""
		if len(prices)>1:
			c = 0
			for price in prices:
				c += 1
				product_price = (((price).strip().rstrip()).encode('utf-8')).replace("\\s+","")
				prod_price.append(product_price)
				if c < len(prices):
					prod_price.append("\n")
		else:
			product_price = (((prices[0]).strip().rstrip()).encode('utf-8')).replace("\\s+","")
			prod_price.append(product_price)
		product_price = ''.join(prod_price)
		
		print "urls::::::::::",response.url
		print "names:::::::::",product_name
		print "prices::::::::",product_price
		"""
		"""
		list2 = []
		for name, price in zip(names,prices):
			list1 = []
			product_name = (((name).strip().rstrip()).encode('utf-8')).replace("\\s+","")
			product_price = (((price).strip().rstrip()).encode('utf-8')).replace("\\s+","")
			print product_name,product_price
			
			for list_part in self.list2:
				if response.url in list_part:
					ref = list_part[0]

			list1.append(ref)
			list1.append(product_name)
		        list1.append(product_price)
			list1.append(response.url)
			list2.append(list1)
		print "URL>>>>>>",response.url
		"""

		for list_part in self.list2:
			if response.url in list_part:
				ref = list_part[0]
		
		if os.path.isfile('../../../media/grofers_rates.csv'):
		        csvfile = open('../../../media/grofers_rates.csv', 'a')
		        fieldnames = ['Reference','Name', 'Price', 'URL']
		        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
		        writer.writerow({'Reference': ref, 'Name': name, 'Price': price, 'URL': response.url})
		 	print "Done"
		
		else:
		        csvfile = open('../../../media/grofers_rates.csv', 'a')
		        fieldnames = ['Reference', 'Name', 'Price', 'URL']
                        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
              	        writer.writeheader()
	                writer.writerow({'Reference': ref, 'Name': name, 'Price': price, 'URL': response.url})
			print "Done"

		

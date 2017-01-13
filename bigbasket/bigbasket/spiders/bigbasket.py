import scrapy
from scrapy.contrib.linkextractors.sgml import SgmlLinkExtractor
from scrapy.selector import Selector
from scrapy.spider import BaseSpider
from xlrd import open_workbook
from openpyxl import Workbook, load_workbook
import xlrd,sys
import openpyxl,os

class linkedInSpider(scrapy.Spider):
	name = "bigbasket"
	allowed_domains = ["bigbasket.com"]

	urls = []
	links = []

	wb = openpyxl.load_workbook('../../../media/Benchmarking.xlsx')
	ws = wb.get_sheet_by_name('Done')
	
	list2 = []
	for row in range(2, ws.max_row + 1):
		list = []
		# Each row in the spreadsheet has data for one census tract.
		ref  = ws['A' + str(row)].value
		link = ws['P' + str(row)].value

		if link and (link.encode('utf8')).startswith("http:"):
			list.append(ref)
			list.append(link)
			list2.append(list) 
	
	#print list2

	for ref,link in list2:
		links.append(link)
	print links
	"""
	wb = openpyxl.load_workbook('data.xlsx')
	first_sheet = wb.get_sheet_names()[0]
	worksheet = wb.get_sheet_by_name(first_sheet)
	
	#here you iterate over the rows in the specific column
	for row in range(2,worksheet.max_row):  
	    for column in "AP":  #Here you can add or reduce the columns
		cell_name = "{}{}".format(column, row)
		#print worksheet[cell_name].value
		if ((worksheet[cell_name].value)):
			urls.append((worksheet[cell_name].value).encode('utf8'))
	print urls
	for url in urls:
	    if (url.encode('utf8')).startswith('http:'):
		links.append(url.encode('utf8'))
	#print links
	"""
	if os.path.isfile('../../../media/bigbasket_rates.csv'):
		os.remove('../../../media/bigbasket_rates.csv')
	
	start_urls = links#("http://www.bigbasket.com/pd/274795/lux-soap-bar-soft-touch-silk-essence-rose-water-150-gm-pouch/",)

	def parse(self, response):
		reload(sys)
		sys.setdefaultencoding('utf-8')
		import base64
		import csv
        	import cStringIO
		sel = Selector(response)
		#hxs = HtmlXPathSelector(text=response.body)
		
		
		names = sel.xpath('//div[@class="uiv2-product-name"]/h1/text()').extract()
		prod_name = []
		prod_price = []
		
		if len(names)>1:
			l = 0
			for name in names:
				l += 1
				product_name = (((name).strip().rstrip()).encode('utf-8')).replace("\n","")
				prod_name.append(product_name)
				if l < len(names):
					prod_name.append("\n")
		else:
			product_name = (((names[0]).strip().rstrip()).encode('utf-8')).replace("\n"," ")
			prod_name.append(product_name)
			product_name = " ".join(product_name.split())
		product_name = (''.join(prod_name)).strip().rstrip()
		product_name = product_name.replace('\t','')
		
			
		
		prices = sel.xpath('//div[@class="uiv2-product-value"]/div[@class="uiv2-price"]/text()').extract()
		
		
		if len(prices)>1:
			c = 0
			for price in prices:
				c += 1
				product_price = (((price).strip().rstrip()).encode('utf-8')).replace("\\s+","")
				prod_price.append(product_price)
				if c < len(prices):
					prod_price.append("\n")
		else:
			product_price = (((prices[0]).strip().rstrip()).encode('utf-8')).replace("\\s+","")
			prod_price.append(product_price)
		product_price = ''.join(prod_price)
		product = []
		print "urls::::::::::",response.url
		print "names:::::::::",product_name
		print "prices::::::::",product_price
		
		"""
		list2 = []
		for name, price in zip(names,prices):
			list1 = []
			product_name = (((name).strip().rstrip()).encode('utf-8')).replace("\\s+","")
			product_price = (((price).strip().rstrip()).encode('utf-8')).replace("\\s+","")
			print product_name,product_price
			
			for list_part in self.list2:
				if response.url in list_part:
					ref = list_part[0]

			list1.append(ref)
			list1.append(product_name)
		        list1.append(product_price)
			list1.append(response.url)
			list2.append(list1)
		print "URL>>>>>>",response.url
		"""
		
		for list_part in self.list2:
			if response.url in list_part:
				ref = list_part[0]
		
		if os.path.isfile('../../../media/bigbasket_rates.csv'):
		        csvfile = open('../../../media/bigbasket_rates.csv', 'a')
		        fieldnames = ['Reference','Name', 'Price', 'URL']
		        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
		        writer.writerow({'Reference': ref, 'Name': product_name, 'Price': product_price, 'URL': response.url})
		 	print "Done"
		
		else:
		        csvfile = open('../../../media/bigbasket_rates.csv', 'a')
		        fieldnames = ['Reference', 'Name', 'Price', 'URL']
                        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
              	        writer.writeheader()
	                writer.writerow({'Reference': ref, 'Name': product_name, 'Price': product_price, 'URL': response.url})
			print "Done"
		
		
